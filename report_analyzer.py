"""
Модуль анализа отчёта по часу продаж.
Читает Excel → генерирует PDF со сравнением подразделений и регионов.
"""
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import os

pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))


def safe_num(v):
    """Преобразует значение ячейки в число, иначе None"""
    if v is None or v == '-':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_report(xlsx_path):
    """Парсит Excel-отчёт и возвращает данные: магазины, подразделения, регионы"""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['Регион']
    
    # Метаданные из шапки
    meta = {'title': '', 'date': '', 'time': ''}
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if row[0]:
            if 'Отчет' in str(row[0]):
                meta['title'] = str(row[0])
            elif 'Дата' in str(row[0]):
                meta['date'] = str(row[0]).replace('Дата формирования: ', '')
            elif 'Время' in str(row[0]):
                meta['time'] = str(row[0]).replace('Время: ', '')
    
    # Находим строку с заголовками колонок
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row[0] == 'Подр':
            header_row_idx = i
            headers = row
            break
    
    if not header_row_idx:
        raise ValueError("Не найдена строка заголовков")
    
    # Индексы важных колонок
    idx = {h: i for i, h in enumerate(headers) if h}
    
    stores = []       # магазины
    totals = []       # итоги по регионам
    
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row[0]:
            continue
        if row[0] == 'Подр':  # ещё одна шапка для итогов
            continue
        
        record = {
            'podr': row[0],
            'shop': row[1],
            'tc': row[2],
            'to': safe_num(row[idx.get('Ср ТО с НДС', 3)]),
            'to_vs_week': safe_num(row[idx.get('Ср. ТО к Нед.', 4)]),
            'to_vs_yesterday': safe_num(row[idx.get('ТО к Вчера', 5)]),
            'plan_pct': safe_num(row[idx.get('План, %', 6)]),
            'kop': safe_num(row[idx.get('КОП', 8)]),
            'pvch': safe_num(row[idx.get('ПвЧ', 11)]),
            'items_check': safe_num(row[idx.get('Штук в чеке', 13)]),
            'sch_shoes': safe_num(row[idx.get('СЧ обувь', 14)]),
            'sch': safe_num(row[idx.get('СЧ', 16)]),
            'traffic': safe_num(row[idx.get('Трафик', 18)]),
            'yui': safe_num(row[idx.get('ЮИ, %', 35)]),
            'silver': safe_num(row[idx.get('Серебро, %', 36)]),
            'gold': safe_num(row[idx.get('Золото, %', 37)]),
            'bags': safe_num(row[idx.get('Сумки,%', 42)]),
        }
        
        if row[0] == 'ИТОГО':
            record['region'] = row[2]
            totals.append(record)
        else:
            stores.append(record)
    
    # Агрегация по подразделениям (средние/медианные)
    podr_agg = {}
    for s in stores:
        p = s['podr']
        if p not in podr_agg:
            podr_agg[p] = {'stores': [], 'name': p}
        podr_agg[p]['stores'].append(s)
    
    for p, data in podr_agg.items():
        stores_list = data['stores']
        data['count'] = len(stores_list)
        for metric in ['to', 'plan_pct', 'kop', 'pvch', 'items_check', 'sch', 'traffic', 'yui', 'silver', 'gold', 'bags']:
            vals = [s[metric] for s in stores_list if s[metric] is not None]
            data[f'avg_{metric}'] = sum(vals) / len(vals) if vals else None
    
    return {
        'meta': meta,
        'stores': stores,
        'totals': totals,
        'podr': podr_agg,
    }


def fmt_num(v, digits=0, percent=False, plus_sign=False):
    """Форматирует число для вывода"""
    if v is None:
        return '—'
    if percent:
        v = v * 100
        s = f"{v:+.{digits}f}%" if plus_sign else f"{v:.{digits}f}%"
        return s
    s = f"{v:+,.{digits}f}" if plus_sign else f"{v:,.{digits}f}"
    return s.replace(',', ' ')


def generate_pdf(data, output_path):
    """Генерирует PDF-отчёт по распарсенным данным"""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A3),
        leftMargin=1*cm, rightMargin=1*cm,
        topMargin=1*cm, bottomMargin=1*cm
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontName='DejaVu-Bold',
                                  fontSize=20, textColor=colors.HexColor('#1F4E78'),
                                  alignment=1, spaceAfter=4)
    subtitle_style = ParagraphStyle('S', parent=styles['Normal'], fontName='DejaVu',
                                     fontSize=10, textColor=colors.HexColor('#555555'),
                                     alignment=1, spaceAfter=8)
    h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontName='DejaVu-Bold',
                        fontSize=14, textColor=colors.HexColor('#1F4E78'),
                        spaceBefore=8, spaceAfter=6)
    normal = ParagraphStyle('N', parent=styles['Normal'], fontName='DejaVu',
                            fontSize=10, leading=12)
    positive = ParagraphStyle('P', parent=styles['Normal'], fontName='DejaVu-Bold',
                              fontSize=10, textColor=colors.HexColor('#005000'), leading=12)
    negative = ParagraphStyle('Neg', parent=styles['Normal'], fontName='DejaVu-Bold',
                              fontSize=10, textColor=colors.HexColor('#B00000'), leading=12)
    
    story = []
    
    # ====== ЗАГОЛОВОК ======
    story.append(Paragraph("Анализ часа продаж: сравнение подразделений и регионов", title_style))
    story.append(Paragraph(
        f"Данные отчёта: {data['meta']['date']} {data['meta']['time']} • "
        f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        subtitle_style
    ))
    
    # Вычисляем kari_total и kz_total заранее, нужно для легенды
    kari_total = next((t for t in data['totals'] if 'Kari' in t.get('region', '')), None)
    kz_total = next((t for t in data['totals'] if 'КЗ' in t.get('region', '')), None)
    
    # Легенда
    kari_plan_str = fmt_num(kari_total.get('plan_pct') if kari_total else None, 1, percent=True)
    kari_kop_str = fmt_num(kari_total.get('kop') if kari_total else None, 1, percent=True)
    kari_sch_str = fmt_num(kari_total.get('sch') if kari_total else None, 0)
    
    legend_text = (
        f"<b>Эталоны:</b> Ср.ТО — лучший по КЗ, План ≥ Kari ({kari_plan_str}), "
        f"КОП ≥ Kari ({kari_kop_str}), ПвЧ ≥ 1.5, Шт/чек ≥ 2.0, СЧ ≥ Kari ({kari_sch_str}), "
        f"Трафик — лучший по КЗ, ЮИ ≥ 10%, Серебро ≥ 5%, Золото ≥ 10%, Сумки ≥ 6%. "
        f"<font backcolor='#7ED97E' color='#003300'><b>&nbsp;Топ-2&nbsp;</b></font> &nbsp;"
        f"<font backcolor='#FF8A8A' color='#660000'><b>&nbsp;Ниже нормы&nbsp;</b></font> &nbsp;"
        f"<font backcolor='#E31B1B' color='#FFFFFF'><b>&nbsp;Ноль&nbsp;</b></font>"
    )
    legend_data = [[Paragraph(legend_text, normal)]]
    legend_tbl = Table(legend_data, colWidths=[27*cm])
    legend_tbl.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(legend_tbl)
    story.append(Spacer(1, 8))
    
    # ====== БЛОК 1: СРАВНЕНИЕ ПОДРАЗДЕЛЕНИЙ (КЗ) ======
    story.append(Paragraph("🏢 Сравнение подразделений (Казахстан)", h2))
    
    podr_sorted = sorted(data['podr'].values(), key=lambda x: x.get('avg_to') or 0, reverse=True)
    
    podr_header = ['Подразделение', 'Магаз.', 'Ср. ТО', 'План %', 'КОП %', 'ПвЧ', 'Шт/чек', 'СЧ', 'Трафик', 'ЮИ %', 'Серебро %', 'Золото %', 'Сумки %']
    podr_rows = [podr_header]
    for p in podr_sorted:
        podr_rows.append([
            p['name'],
            str(p['count']),
            fmt_num(p['avg_to'], 0),
            fmt_num(p['avg_plan_pct'], 1, percent=True),
            fmt_num(p['avg_kop'], 1, percent=True),
            fmt_num(p['avg_pvch'], 2),
            fmt_num(p['avg_items_check'], 2),
            fmt_num(p['avg_sch'], 0),
            fmt_num(p['avg_traffic'], 0),
            fmt_num(p['avg_yui'], 1, percent=True),
            fmt_num(p['avg_silver'], 1, percent=True),
            fmt_num(p['avg_gold'], 1, percent=True),
            fmt_num(p['avg_bags'], 1, percent=True),
        ])
    
    tbl = Table(podr_rows, repeatRows=1, hAlign='CENTER')
    ts = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'DejaVu'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVu-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])
    
    # Эталоны — зафиксированы пользователем (kari_total/kz_total определены выше)
    GREEN = colors.HexColor('#7ED97E')        # насыщенный зелёный
    RED = colors.HexColor('#FF8A8A')          # насыщенный красный
    RED_AGGRESSIVE = colors.HexColor('#E31B1B')  # агрессивный красный для нулей
    GREEN_TEXT = colors.HexColor('#003300')
    RED_TEXT = colors.HexColor('#660000')
    WHITE = colors.white
    
    # Лучшие значения по КЗ (для Ср.ТО и Трафика)
    best_to_kz = max((p.get('avg_to') for p in data['podr'].values() if p.get('avg_to') is not None), default=None)
    best_traffic_kz = max((p.get('avg_traffic') for p in data['podr'].values() if p.get('avg_traffic') is not None), default=None)
    
    # Пороги (эталоны). None = не проверяем
    kari_plan = kari_total.get('plan_pct') if kari_total else None
    kari_kop = kari_total.get('kop') if kari_total else None
    kari_sch = kari_total.get('sch') if kari_total else None
    
    # Правила: ключ → (эталон, правило сравнения)
    # Правила:
    #   'eq_best' — равен лучшему по КЗ (для ТО и Трафика)
    #   'gte' — значение >= эталона
    # Проценты хранятся как доли (0.10 = 10%)
    RULES = {
        'to':          ('eq_best', best_to_kz),
        'plan_pct':    ('gte',     kari_plan),   # выше Kari
        'kop':         ('gte',     kari_kop),    # выше Kari
        'pvch':        ('gte',     1.5),
        'items_check': ('gte',     2.0),
        'sch':         ('gte',     kari_sch),    # не ниже Kari
        'traffic':     ('eq_best', best_traffic_kz),
        'yui':         ('gte',     0.10),   # не менее 10%
        'silver':      ('gte',     0.05),   # не ниже 5%
        'gold':        ('gte',     0.10),   # не ниже 10%
        'bags':        ('gte',     0.06),   # не ниже 6%
    }
    
    def paint_cell(style_obj, col, row, value, rule_type, target):
        """Красит только красным (не дотягивает норму) или ярко-красным (ноль).
        Зелёная подсветка применяется отдельно — только для топ-2 лучших."""
        if value is None:
            return
        # Нули — агрессивный красный
        if value == 0:
            style_obj.add('BACKGROUND', (col, row), (col, row), RED_AGGRESSIVE)
            style_obj.add('TEXTCOLOR', (col, row), (col, row), WHITE)
            style_obj.add('FONTNAME', (col, row), (col, row), 'DejaVu-Bold')
            return
        if target is None:
            return
        if rule_type == 'eq_best':
            passed = abs(value - target) < 1e-6
        elif rule_type == 'gte':
            passed = value >= target
        else:
            return
        # Норма не выполнена → красный; выполнена → без заливки (зелёный ставим отдельно)
        if not passed:
            style_obj.add('BACKGROUND', (col, row), (col, row), RED)
            style_obj.add('TEXTCOLOR', (col, row), (col, row), RED_TEXT)
    
    def highlight_top2(style_obj, col, rows_with_values):
        """Зелёным красит 2 самых высоких значения в колонке (нули игнорируем)."""
        valid = [(idx, v) for idx, v in rows_with_values if v is not None and v != 0]
        if len(valid) < 2:
            return
        sorted_desc = sorted(valid, key=lambda x: x[1], reverse=True)
        for idx, _ in sorted_desc[:2]:
            style_obj.add('BACKGROUND', (col, idx), (col, idx), GREEN)
            style_obj.add('TEXTCOLOR', (col, idx), (col, idx), GREEN_TEXT)
            style_obj.add('FONTNAME', (col, idx), (col, idx), 'DejaVu-Bold')
    
    # Колонки таблицы подразделений:
    # 0=Подр, 1=Магаз.кол-во, 2=Ср.ТО, 3=План%, 4=КОП%, 5=ПвЧ, 6=Шт/чек, 7=СЧ, 8=Трафик, 9=ЮИ%, 10=Серебро%, 11=Золото%, 12=Сумки%
    metrics_map = [
        (2, 'to'), (3, 'plan_pct'), (4, 'kop'), (5, 'pvch'),
        (6, 'items_check'), (7, 'sch'), (8, 'traffic'),
        (9, 'yui'), (10, 'silver'), (11, 'gold'), (12, 'bags'),
    ]
    # Сначала красим красным всё что не дотягивает норму
    for i, p in enumerate(podr_sorted, 1):
        for col, key in metrics_map:
            rule_type, target = RULES.get(key, (None, None))
            paint_cell(ts, col, i, p.get(f'avg_{key}'), rule_type, target)
    # Потом сверху накрываем зелёным топ-2 лучших (перекрывает красный если случайно совпало)
    for col, key in metrics_map:
        rows_vals = [(i, p.get(f'avg_{key}')) for i, p in enumerate(podr_sorted, 1)]
        highlight_top2(ts, col, rows_vals)
    
    tbl.setStyle(ts)
    story.append(tbl)
    story.append(Spacer(1, 6))
    
    # ====== БЛОК 2: СРАВНЕНИЕ РЕГИОНОВ ======
    story.append(Paragraph("🌍 Сравнение регионов", h2))
    
    regions_sorted = sorted(data['totals'], key=lambda x: x.get('to') or 0, reverse=True)
    
    reg_header = ['Регион', 'Ср. ТО', 'План %', 'КОП %', 'ПвЧ', 'Шт/чек', 'СЧ', 'Трафик', 'ЮИ %', 'Серебро %', 'Золото %', 'Сумки %']
    reg_rows = [reg_header]
    for r in regions_sorted:
        reg_rows.append([
            r.get('region', '').replace('Итого по ', ''),
            fmt_num(r['to'], 0),
            fmt_num(r['plan_pct'], 1, percent=True),
            fmt_num(r['kop'], 1, percent=True),
            fmt_num(r['pvch'], 2),
            fmt_num(r['items_check'], 2),
            fmt_num(r['sch'], 0),
            fmt_num(r['traffic'], 0),
            fmt_num(r['yui'], 1, percent=True),
            fmt_num(r['silver'], 1, percent=True),
            fmt_num(r['gold'], 1, percent=True),
            fmt_num(r['bags'], 1, percent=True),
        ])
    
    tbl2 = Table(reg_rows, repeatRows=1, hAlign='CENTER')
    ts2 = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'DejaVu'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVu-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])
    
    # Подсветка регионов по тем же правилам
    # Для ТО и Трафика в таблице регионов "лучший" = максимум среди регионов (исключая Kari)
    comparable_totals = [r for r in regions_sorted if 'Kari' not in r.get('region', '')]
    best_to_reg = max((r.get('to') for r in comparable_totals if r.get('to') is not None), default=None)
    best_traffic_reg = max((r.get('traffic') for r in comparable_totals if r.get('traffic') is not None), default=None)
    
    REGION_RULES = {
        'to':          ('eq_best', best_to_reg),
        'plan_pct':    ('gte',     kari_plan),
        'kop':         ('gte',     kari_kop),
        'pvch':        ('gte',     1.5),
        'items_check': ('gte',     2.0),
        'sch':         ('gte',     kari_sch),
        'traffic':     ('eq_best', best_traffic_reg),
        'yui':         ('gte',     0.10),
        'silver':      ('gte',     0.05),
        'gold':        ('gte',     0.10),
        'bags':        ('gte',     0.06),
    }
    
    # Колонки: 0=Регион, 1=ТО, 2=План, 3=КОП, 4=ПвЧ, 5=Шт/чек, 6=СЧ, 7=Трафик, 8=ЮИ, 9=Серебро, 10=Золото, 11=Сумки
    reg_metrics = [
        (1, 'to'), (2, 'plan_pct'), (3, 'kop'), (4, 'pvch'),
        (5, 'items_check'), (6, 'sch'), (7, 'traffic'),
        (8, 'yui'), (9, 'silver'), (10, 'gold'), (11, 'bags'),
    ]
    
    # Индексы строк регионов без Kari (Kari — эталон, его не красим)
    comparable_rows = [(i, r) for i, r in enumerate(regions_sorted, 1) if 'Kari' not in r.get('region', '')]
    
    # Сначала красный для не-норм
    for i, r in enumerate(regions_sorted, 1):
        if 'Kari' in r.get('region', ''):
            ts2.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#E8E8E8'))
            ts2.add('FONTNAME', (0, i), (-1, i), 'DejaVu-Bold')
            continue
        for col, key in reg_metrics:
            rule_type, target = REGION_RULES.get(key, (None, None))
            paint_cell(ts2, col, i, r.get(key), rule_type, target)
        # КЗ — рамка
        if 'КЗ' in r.get('region', ''):
            ts2.add('LINEABOVE', (0, i), (-1, i), 1.5, colors.HexColor('#1F4E78'))
            ts2.add('LINEBELOW', (0, i), (-1, i), 1.5, colors.HexColor('#1F4E78'))
            ts2.add('FONTNAME', (0, i), (0, i), 'DejaVu-Bold')
    
    # Потом зелёным топ-2 (без Kari)
    for col, key in reg_metrics:
        rows_vals = [(i, r.get(key)) for i, r in comparable_rows]
        highlight_top2(ts2, col, rows_vals)
    
    tbl2.setStyle(ts2)
    story.append(tbl2)
    
    story.append(PageBreak())
    
    # ====== БЛОК 3: АНАЛИЗ ОТКЛОНЕНИЙ ПОДРАЗДЕЛЕНИЙ ======
    story.append(Paragraph("🔍 Анализ по подразделениям", h2))
    
    if kari_total:
        story.append(Paragraph(
            f"<i>Эталоны: ТО — лучший по КЗ, план ≥ {fmt_num(kari_total.get('plan_pct'), 1, percent=True)} (Kari), "
            f"КОП ≥ {fmt_num(kari_total.get('kop'), 1, percent=True)} (Kari), ПвЧ ≥ 1.5, шт/чек ≥ 2.0, "
            f"СЧ ≥ {fmt_num(kari_total.get('sch'), 0)} (Kari), трафик — лучший по КЗ, "
            f"ЮИ ≥ 10%, серебро ≥ 5%, золото ≥ 10%, сумки ≥ 6%</i>",
            normal
        ))
        story.append(Spacer(1, 8))
        
        # Проверка метрики: возвращает (passed, value_str, target_str)
        def check_metric(value, rule_type, target, fmt_percent=False, digits=1):
            if value is None:
                return None, '—', '—'
            if value == 0:
                return False, fmt_num(value, digits, percent=fmt_percent) + ' ⚠️', fmt_num(target, digits, percent=fmt_percent) if target else '—'
            if target is None:
                return None, fmt_num(value, digits, percent=fmt_percent), '—'
            if rule_type == 'eq_best':
                passed = abs(value - target) < 1e-6
            else:  # gte
                passed = value >= target
            return passed, fmt_num(value, digits, percent=fmt_percent), fmt_num(target, digits, percent=fmt_percent)
        
        # Для каждого подразделения делаем карточку
        # Генерируем карточки для всех подразделений
        all_cards = []
        for p in podr_sorted:
            checks = [
                ('ТО',       p.get('avg_to'),          'eq_best', best_to_kz,     False, 0),
                ('План',     p.get('avg_plan_pct'),    'gte',     kari_plan,      True,  1),
                ('КОП',      p.get('avg_kop'),         'gte',     kari_kop,       True,  1),
                ('ПвЧ',      p.get('avg_pvch'),        'gte',     1.5,            False, 2),
                ('Шт/чек',   p.get('avg_items_check'), 'gte',     2.0,            False, 2),
                ('СЧ',       p.get('avg_sch'),         'gte',     kari_sch,       False, 0),
                ('Трафик',   p.get('avg_traffic'),     'eq_best', best_traffic_kz,False, 0),
                ('ЮИ',       p.get('avg_yui'),         'gte',     0.10,           True,  1),
                ('Серебро',  p.get('avg_silver'),      'gte',     0.05,           True,  1),
                ('Золото',   p.get('avg_gold'),        'gte',     0.10,           True,  1),
                ('Сумки',    p.get('avg_bags'),        'gte',     0.06,           True,  1),
            ]
            
            results = []
            passed_count = 0
            failed_count = 0
            for label, val, rule, tgt, is_pct, dig in checks:
                passed, v_str, t_str = check_metric(val, rule, tgt, is_pct, dig)
                if passed is True:
                    passed_count += 1
                elif passed is False:
                    failed_count += 1
                results.append((label, passed, v_str, t_str, val))
            
            total_checks = passed_count + failed_count
            score_pct = (passed_count / total_checks * 100) if total_checks else 0
            
            # Статус-цвета — насыщенные
            if score_pct >= 70:
                status_bg = colors.HexColor('#2E7D32')  # тёмно-зелёный
                status_emoji = '🟢'
                status_text = 'ХОРОШО'
            elif score_pct >= 40:
                status_bg = colors.HexColor('#E65100')  # тёмно-оранжевый
                status_emoji = '🟡'
                status_text = 'СРЕДНЕ'
            else:
                status_bg = colors.HexColor('#B71C1C')  # тёмно-красный
                status_emoji = '🔴'
                status_text = 'СЛАБО'
            
            # Шапка карточки — компактная
            header_data = [[
                Paragraph(
                    f"<font color='#FFFFFF' size='12'><b>{p['name']}</b></font> "
                    f"<font color='#FFFFFF' size='9'>· {p['count']} маг · {fmt_num(p['avg_to'])} тг</font>",
                    normal
                ),
                Paragraph(
                    f"<font color='#FFFFFF' size='12'><b>{status_emoji} {status_text} {passed_count}/{total_checks}</b></font>",
                    ParagraphStyle('s', parent=normal, alignment=2)
                )
            ]]
            header_tbl = Table(header_data, colWidths=[12*cm, 7.5*cm])
            header_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), status_bg),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            # Таблица метрик
            metrics_rows = []
            for label, passed, v_str, t_str, raw_val in results:
                if raw_val == 0:
                    val_color = '#FFFFFF'
                    label_color = '#FFFFFF'
                    target_color = '#FFE0E0'
                    icon = '⚠'
                elif passed is True:
                    val_color = '#1B5E20'
                    label_color = '#000000'
                    target_color = '#757575'
                    icon = '✓'
                elif passed is False:
                    val_color = '#B71C1C'
                    label_color = '#000000'
                    target_color = '#757575'
                    icon = '✗'
                else:
                    val_color = '#333333'
                    label_color = '#333333'
                    target_color = '#999999'
                    icon = '·'
                
                metrics_rows.append([
                    Paragraph(f"<font size='10' color='{label_color}'><b>{icon} {label}</b></font>", normal),
                    Paragraph(f"<font size='11' color='{val_color}'><b>{v_str}</b></font>", normal),
                    Paragraph(f"<font size='9' color='{target_color}'>эталон {t_str}</font>", normal),
                ])
            
            mt = Table(metrics_rows, colWidths=[5*cm, 4*cm, 10.5*cm])
            mt_style = TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'DejaVu'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#888888')),
            ])
            for i, (label, passed, v_str, t_str, raw_val) in enumerate(results):
                if raw_val == 0:
                    mt_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#E31B1B'))
                elif passed is True:
                    mt_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#C8E6C9'))
                elif passed is False:
                    mt_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFCDD2'))
            mt.setStyle(mt_style)
            
            # Карточка = шапка + таблица метрик, обёрнутые в ячейку
            card = Table([[header_tbl], [mt]], colWidths=[19.5*cm])
            card.setStyle(TableStyle([
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            all_cards.append(card)
        
        # Расставляем карточки 2-в-ряд
        rows_2x = []
        for i in range(0, len(all_cards), 2):
            left = all_cards[i]
            right = all_cards[i+1] if i+1 < len(all_cards) else ''
            rows_2x.append([left, right])
        
        grid = Table(rows_2x, colWidths=[20*cm, 20*cm])
        grid.setStyle(TableStyle([
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(grid)
    
    doc.build(story)
    return output_path


if __name__ == '__main__':
    # Демо: запуск на вашем файле
    input_file = '/mnt/user-data/uploads/Отчет_по_часу_продаж_3_2_Регион__shoes__-_2026-04-16T154228_674.xlsx'
    output_file = '/home/claude/анализ_час_продаж.pdf'
    
    data = parse_report(input_file)
    print(f"✅ Распаршено: {len(data['stores'])} магазинов, {len(data['podr'])} подразделений, {len(data['totals'])} регионов")
    
    generate_pdf(data, output_file)
    print(f"✅ PDF создан: {output_file}")
